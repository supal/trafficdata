#!/usr/bin/env python3
"""
A MCP server with tools for comprehensive traffic data analysis.
Supports analysis of all vehicle type combinations.
"""

import asyncio
from pathlib import Path
from mcp.server import Server
import mcp.types as types
from openpyxl import load_workbook
import statistics
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
import base64
from io import BytesIO


# Create the MCP server
server = Server("traffic-mcp-server")

# Column mappings for vehicle types
VEHICLE_TYPES = {
    "all_vehicles": {"count_col": 2, "speed_col": 3, "name": "All Vehicles"},
    "heavy_vehicles": {"count_col": 4, "speed_col": 5, "name": "Heavy Vehicles"},
    "passenger_cars": {"count_col": 6, "speed_col": 7, "name": "Passenger Cars"},
    "heavy_with_trailer": {"count_col": 8, "speed_col": 9, "name": "Heavy Vehicles with Trailer"},
    "heavy_without_trailer": {"count_col": 10, "speed_col": 11, "name": "Heavy Vehicles without Trailer"},
    "three_axle_with_trailer": {"count_col": 12, "speed_col": 13, "name": "Three-Axle Tractor with Trailer"},
    "two_axle_with_trailer": {"count_col": 14, "speed_col": 15, "name": "Two-Axle Tractor with Trailer"},
    "three_axle_without_trailer": {"count_col": 16, "speed_col": 17, "name": "Three-Axle Tractor without Trailer"},
    "two_axle_without_trailer": {"count_col": 18, "speed_col": 19, "name": "Two-Axle Tractor without Trailer"},
    "passenger_with_trailer": {"count_col": 20, "speed_col": 21, "name": "Passenger Cars with Trailer"},
    "passenger_without_trailer": {"count_col": 22, "speed_col": 23, "name": "Passenger Cars without Trailer"},
}


def load_traffic_data():
    """Load all traffic data from Excel file."""
    excel_file = Path(__file__).parent / "data" / "trafikverket_data_9dccacb0.xlsx"
    if not excel_file.exists():
        return None
    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        time_col = 1
        timestamps = []
        data = {key: [] for key in VEHICLE_TYPES.keys()}
        
        for row_idx in range(3, ws.max_row + 1):
            # Get timestamp
            time_cell = ws.cell(row=row_idx, column=time_col).value
            timestamp = None
            if time_cell:
                try:
                    if isinstance(time_cell, datetime):
                        timestamp = time_cell
                    else:
                        timestamp = datetime.fromisoformat(str(time_cell))
                except:
                    pass
            
            if not timestamp:
                continue
            
            # Extract speeds for all vehicle types
            row_has_data = False
            vehicle_speeds = {}
            
            for vehicle_key, col_info in VEHICLE_TYPES.items():
                speed_col = col_info["speed_col"]
                speed_cell = ws.cell(row=row_idx, column=speed_col).value
                
                speed = None
                if speed_cell and str(speed_cell).strip() and str(speed_cell) != "0,0":
                    try:
                        speed = float(str(speed_cell).replace(",", "."))
                        if speed > 0:
                            vehicle_speeds[vehicle_key] = speed
                            row_has_data = True
                    except:
                        pass
            
            # Add row if we have at least one speed value
            if row_has_data:
                timestamps.append(timestamp)
                for vehicle_key in VEHICLE_TYPES.keys():
                    data[vehicle_key].append(vehicle_speeds.get(vehicle_key, 0))
        
        data["timestamps"] = timestamps
        return data
    except Exception as e:
        return {"error": str(e)}


def calculate_statistics(speeds):
    """Calculate statistics for a speed list."""
    valid_speeds = [s for s in speeds if s > 0]
    if not valid_speeds:
        return None
    
    return {
        "count": len(valid_speeds),
        "average": statistics.mean(valid_speeds),
        "min": min(valid_speeds),
        "max": max(valid_speeds),
        "median": statistics.median(valid_speeds),
    }


def format_statistics(vehicle_name, stats):
    """Format statistics for display."""
    if not stats:
        return f"{vehicle_name}:\n  No data available"
    
    result = f"{vehicle_name}:\n"
    result += f"  Average Speed: {stats['average']:.2f} km/h\n"
    result += f"  Min Speed: {stats['min']:.2f} km/h\n"
    result += f"  Max Speed: {stats['max']:.2f} km/h\n"
    result += f"  Median Speed: {stats['median']:.2f} km/h\n"
    result += f"  Data Points: {stats['count']}"
    return result


def generate_speed_graph(vehicle_types=None):
    """Generate a graph showing average speed over time for selected vehicle types."""
    try:
        data = load_traffic_data()
        if not data or "error" in data:
            return None
        
        timestamps = data.get("timestamps", [])
        if not timestamps:
            return None
        
        # Default to main vehicle types if not specified
        if vehicle_types is None:
            vehicle_types = ["heavy_vehicles", "passenger_cars"]
        
        # Create figure and axis
        fig, ax = plt.subplots(figsize=(14, 7))
        
        # Color map for different vehicle types
        colors = {
            "heavy_vehicles": "red",
            "passenger_cars": "blue",
            "heavy_with_trailer": "darkred",
            "heavy_without_trailer": "lightcoral",
            "passenger_with_trailer": "darkblue",
            "passenger_without_trailer": "lightblue",
            "two_axle_with_trailer": "orange",
            "three_axle_with_trailer": "darkorange",
            "two_axle_without_trailer": "yellow",
            "three_axle_without_trailer": "gold",
        }
        
        # Plot each vehicle type
        for vehicle_key in vehicle_types:
            if vehicle_key in VEHICLE_TYPES and vehicle_key in data:
                speeds = data[vehicle_key]
                label = VEHICLE_TYPES[vehicle_key]["name"]
                color = colors.get(vehicle_key, "black")
                ax.plot(timestamps, speeds, label=label, color=color, linewidth=2, marker='o', markersize=3, alpha=0.7)
        
        # Format the plot
        ax.set_xlabel('Time', fontsize=12)
        ax.set_ylabel('Average Speed (km/h)', fontsize=12)
        ax.set_title('Average Speed Over Time by Vehicle Type', fontsize=14, fontweight='bold')
        ax.legend(fontsize=9, loc='best')
        ax.grid(True, alpha=0.3)
        
        # Format x-axis to show dates nicely
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d %H:%M'))
        fig.autofmt_xdate(rotation=45, ha='right')
        
        # Tight layout to prevent label cutoff
        fig.tight_layout()
        
        # Save to bytes buffer
        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=100)
        buf.seek(0)
        
        # Encode to base64
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)
        
        return img_base64
    except Exception as e:
        return None


@server.list_tools()
async def list_tools() -> list[types.Tool]:
    """List all available tools."""
    tools = [
        types.Tool(
            name="ping",
            description="Returns 'pong' as a simple test",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": [],
            },
        ),
        types.Tool(
            name="sum",
            description="Adds two numbers together",
            inputSchema={
                "type": "object",
                "properties": {
                    "a": {"type": "number", "description": "First number"},
                    "b": {"type": "number", "description": "Second number"},
                },
                "required": ["a", "b"],
            },
        ),
        types.Tool(
            name="read_file",
            description="Reads and returns the contents of a text file",
            inputSchema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to the file to read"},
                },
                "required": ["path"],
            },
        ),
    ]
    
    # Add vehicle type tools
    for vehicle_key, vehicle_info in VEHICLE_TYPES.items():
        tools.append(
            types.Tool(
                name=f"average_speed_{vehicle_key}",
                description=f"Returns the average speed of {vehicle_info['name']} from traffic data",
                inputSchema={
                    "type": "object",
                    "properties": {},
                    "required": [],
                },
            )
        )
    
    # Add graph tools
    tools.extend([
        types.Tool(
            name="speed_graph",
            description="Generates a graph showing average speed over time (default: heavy vehicles and passenger cars)",
            inputSchema={
                "type": "object",
                "properties": {
                    "vehicle_types": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of vehicle types to include (optional)",
                    },
                },
                "required": [],
            },
        ),
        types.Tool(
            name="all_vehicle_statistics",
            description="Returns average speed statistics for all vehicle types",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": [],
            },
        ),
    ])
    
    return tools


@server.call_tool()
async def call_tool(name: str, arguments: dict) -> list[types.TextContent | types.ImageContent]:
    """Execute a tool by name with the given arguments."""
    if name == "ping":
        return [types.TextContent(type="text", text="pong")]
    
    elif name == "sum":
        a = arguments.get("a")
        b = arguments.get("b")
        if a is None or b is None:
            return [types.TextContent(type="text", text="Error: Missing 'a' or 'b' argument")]
        result = a + b
        return [types.TextContent(type="text", text=f"{result}")]
    
    elif name == "read_file":
        path = arguments.get("path")
        if not path:
            return [types.TextContent(type="text", text="Error: Missing 'path' argument")]
        try:
            file_path = Path(path)
            if not file_path.exists():
                return [types.TextContent(type="text", text=f"Error: File not found: {path}")]
            contents = file_path.read_text()
            return [types.TextContent(type="text", text=contents)]
        except Exception as e:
            return [types.TextContent(type="text", text=f"Error reading file: {str(e)}")]
    
    elif name.startswith("average_speed_"):
        # Extract vehicle type from tool name
        vehicle_key = name[len("average_speed_"):]
        
        data = load_traffic_data()
        if not data or "error" in data:
            return [types.TextContent(type="text", text=f"Error loading traffic data: {data.get('error', 'Unknown error')}")]
        
        if vehicle_key not in VEHICLE_TYPES:
            return [types.TextContent(type="text", text=f"Unknown vehicle type: {vehicle_key}")]
        
        speeds = data.get(vehicle_key, [])
        vehicle_name = VEHICLE_TYPES[vehicle_key]["name"]
        
        stats = calculate_statistics(speeds)
        result = format_statistics(vehicle_name, stats)
        
        return [types.TextContent(type="text", text=result)]
    
    elif name == "all_vehicle_statistics":
        data = load_traffic_data()
        if not data or "error" in data:
            return [types.TextContent(type="text", text=f"Error loading traffic data: {data.get('error', 'Unknown error')}")]
        
        result = "Average Speed Statistics for All Vehicle Types:\n" + "=" * 70 + "\n\n"
        
        for vehicle_key, vehicle_info in VEHICLE_TYPES.items():
            speeds = data.get(vehicle_key, [])
            stats = calculate_statistics(speeds)
            if stats:
                result += format_statistics(vehicle_info["name"], stats)
                result += "\n\n"
        
        return [types.TextContent(type="text", text=result)]
    
    elif name == "speed_graph":
        # Get vehicle types from arguments if provided
        vehicle_types = arguments.get("vehicle_types")
        
        img_base64 = generate_speed_graph(vehicle_types)
        if not img_base64:
            return [types.TextContent(type="text", text="Error generating graph")]
        
        return [types.ImageContent(type="image", data=img_base64, mimeType="image/png")]
    
    else:
        return [types.TextContent(type="text", text=f"Unknown tool: {name}")]


async def main():
    """Run the MCP server over stdio."""
    from mcp.server.stdio import stdio_server
    
    async with stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            server.create_initialization_options()
        )
if __name__ == "__main__":
    asyncio.run(main())
